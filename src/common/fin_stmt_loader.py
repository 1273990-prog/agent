import sys
import os
import io
import json
from typing import Dict, Any, Optional
from datetime import date
from decimal import Decimal

# Windows CP949 콘솔 유니코드 출력 처리
if hasattr(sys.stdout, 'buffer'):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Add 'src' directory to Python search path if not already present
current_dir = os.path.dirname(os.path.abspath(__file__))  # src/common
src_dir = os.path.abspath(os.path.join(current_dir, '..'))  # src
project_root = os.path.abspath(os.path.join(src_dir, '..')) # agent (src와 같은 레벨/프로젝트 루트)

if src_dir not in sys.path:
    sys.path.insert(0, src_dir)

import services.opendart_service
from services.service_factory import ServiceFactory
from common.constants import AgentConstants
from common.utils import AgentUtils
from common.db_conn import DbConn


def get_default_years():
    """
    최근 10년 수집 연도를 동적으로 계산합니다.
    DART 정기 공시 특성을 고려하여 전년도(current_year - 1) 기준 10년을 기본값으로 설정합니다.
    예: 2026년 기준 → 시작연도 2016, 종료연도 2025
    """
    current_year = date.today().year
    end_year = current_year - 1
    start_year = end_year - 9
    return start_year, end_year


def clean_decimal_amount(val_str: Any) -> str:
    """
    OpenDART 수치 문자열을 Decimal 정수/소수 수치 문자열로 변환합니다. (numeric 컬럼 저장용)
    """
    if val_str is None:
        return "0"
    cleaned = str(val_str).replace(",", "").strip()
    if not cleaned or cleaned in ("-", "None", "null"):
        return "0"
    try:
        return str(Decimal(cleaned))
    except Exception:
        return "0"


def clean_int_ord(val_str: Any) -> int:
    """
    OpenDART 'ord' 표시 순서 문자열을 int 정수로 변환합니다. (integer 컬럼 저장용)
    """
    if val_str is None:
        return 0
    cleaned = str(val_str).strip()
    if not cleaned or not cleaned.isdigit():
        return 0
    try:
        return int(cleaned)
    except Exception:
        return 0


def save_company_financial_statements_to_db(
    corp_name: str, 
    corp_code: Optional[str] = None, 
    start_year: Optional[int] = None, 
    end_year: Optional[int] = None
) -> bool:
    """
    OpendartService를 통해 지정 기업의 기간별 정기보고서(사업보고서) 재무제표를 조회하고,
    DbConn 및 AgentUtils.get_rule_no()를 활용하여 DB의 'fin_stmt_info' 테이블에 적재합니다.
    """
    def_start, def_end = get_default_years()
    s_year = start_year if start_year is not None else def_start
    e_year = end_year if end_year is not None else def_end

    config = AgentUtils.load_config("agent_key.json")
    if not isinstance(config, dict) or not config:
        print("[오류] agent_key.json 설정 정보를 로드할 수 없습니다.")
        return False

    crtfc_key = config.get("dart_crtfc_key", "")
    if not crtfc_key:
        print("[오류] OpenDART 인증키(dart_crtfc_key)가 설정되어 있지 않습니다.")
        return False

    opendart_service = ServiceFactory.create(AgentConstants.OPENDART)
    if not opendart_service:
        print("[오류] OPENDART 서비스를 초기화할 수 없습니다.")
        return False

    # 1. DbConn을 통해 'corp_code_info'에서 기업의 PK (rule_no) 및 DART corp_code 조회
    db_net_value_select = {
        "host": config.get("db_host", ""),
        "port": int(config.get("port", 5432)),
        "database": config.get("database", ""),
        "username": config.get("username", ""),
        "password": config.get("password", ""),
        "action": AgentConstants.SELECT
    }

    try:
        db_conn_select = DbConn(json.dumps(db_net_value_select))
        select_req = {
            "query_key": "SELECT_CORP_CODE_BY_NAME",
            "params": {"corp_name": corp_name}
        }
        db_conn_select.create_request(json.dumps(select_req))
        select_res_str = db_conn_select.create_response()
        select_rows = json.loads(select_res_str)

        if not select_rows or not select_rows[0].get("rule_no"):
            print(f"[오류] corp_code_info에서 '{corp_name}'의 PK (rule_no)를 찾을 수 없습니다.")
            return False

        fk_rule_no = select_rows[0]["rule_no"]
        db_corp_code = select_rows[0].get("corp_code", "")
        target_corp_code = corp_code or db_corp_code

        print(f"  [확인] '{corp_name}' FK (corp_code_info.rule_no): {fk_rule_no}, DART corp_code: {target_corp_code}")

    except Exception as e:
        print(f"[오류] DB에서 '{corp_name}' FK 조회 실패: {e}")
        return False

    # 2. 기존 DB 데이터 중 해당 기업(corp_no)의 fin_stmt_info 데이터 삭제 (중복 저장 방지)
    db_net_value_delete = {
        "host": config.get("db_host", ""),
        "port": int(config.get("port", 5432)),
        "database": config.get("database", ""),
        "username": config.get("username", ""),
        "password": config.get("password", ""),
        "action": AgentConstants.DELETE
    }
    try:
        db_conn_delete = DbConn(json.dumps(db_net_value_delete))
        delete_req = {
            "query_key": "DELETE_FIN_STMT_INFO_BY_CORP_NO",
            "params": {"corp_no": fk_rule_no}
        }
        db_conn_delete.create_request(json.dumps(delete_req))
        del_res_str = db_conn_delete.create_response()
        print(f"  [기존 데이터 삭제] '{corp_name}' 기존 fin_stmt_info 레코드 정리 완료: {del_res_str}")
    except Exception as e:
        print(f"[경고] 기존 데이터 삭제 처리 실패: {e}")

    # 3. OpenDART 정기보고서(사업보고서 11011) 데이터 수집 및 bsns_year + account_nm 중복 방지
    print("\n" + "=" * 95)
    print(f"        [ {corp_name} {s_year}~{e_year}년 ({e_year - s_year + 1}년치) 재무제표 DB(fin_stmt_info) 중복 제거 적재 ]")
    print("=" * 95)

    reprt_code = "11011"
    all_insert_params = []
    all_summary = {}
    seen_keys = set()  # (bsns_year, account_nm) 중복 방지 집합

    for year in range(s_year, e_year + 1):
        target_fs_div = "CFS"
        price_request = {
            "crtfc_key": crtfc_key,
            "corp_code": target_corp_code,
            "bsns_year": str(year),
            "reprt_code": reprt_code,
            "fs_div": target_fs_div
        }

        try:
            res_json = opendart_service.get_financial_statements_to_json(json.dumps(price_request))
            status = res_json.get("status")
            message = res_json.get("message")

            # CFS 결과가 없거나 실패 시 OFS로 재시도
            if status != "000" or not res_json.get("list"):
                target_fs_div = "OFS"
                price_request["fs_div"] = target_fs_div
                res_json = opendart_service.get_financial_statements_to_json(json.dumps(price_request))
                status = res_json.get("status")
                message = res_json.get("message")

            if status == "000":
                items = res_json.get("list", [])
                summary = {}
                added_count = 0

                for item in items:
                    bsns_year = str(item.get("bsns_year", "")).strip()
                    account_nm = str(item.get("account_nm", "")).strip()
                    dedup_key = (bsns_year, account_nm)

                    if dedup_key in seen_keys:
                        continue
                    seen_keys.add(dedup_key)

                    dec_amount = clean_decimal_amount(item.get("thstrm_amount", "0"))

                    param = {
                        "rule_no": AgentUtils.get_rule_no(),
                        "corp_no": fk_rule_no,
                        "bsns_year": bsns_year,
                        "account_id": str(item.get("account_id", "")),
                        "account_nm": account_nm,
                        "account_detail": str(item.get("account_detail", "")),
                        "thstrm_amount": dec_amount,
                        "ord": clean_int_ord(item.get("ord", 0)),
                        "thstrm_nm": str(item.get("thstrm_nm", "")),
                        "fs_div": str(item.get("fs_div", target_fs_div)),
                        "fs_nm": str(item.get("fs_nm", "연결재무제표" if target_fs_div == "CFS" else "재무제표")),
                        "sj_div": str(item.get("sj_div", "")),
                        "sj_nm": str(item.get("sj_nm", "")),
                        "rcept_no": str(item.get("rcept_no", "")),
                        "reprt_code": str(item.get("reprt_code", ""))
                    }
                    all_insert_params.append(param)
                    added_count += 1

                    summary[account_nm] = dec_amount

                all_summary[year] = summary
                print(f"  [성공] {year}년 사업보고서 ({target_fs_div}: 원본 {len(items)}개 중 bsns_year+account_nm 중복제거 후 {added_count}개 반영)")
            else:
                print(f"  [실패] {year}년 사업보고서 조회 실패 (상태코드: {status}, 메시지: {message})")

        except Exception as e:
            print(f"  [오류] {year}년 사업보고서 수집 중 예외 발생: {e}")

    # 4. DbConn을 활용하여 fin_stmt_info에 바인딩 데이터 일괄 삽입
    if not all_insert_params:
        print("[경고] 저장할 재무제표 데이터가 없습니다.")
        return False

    db_net_value_insert = {
        "host": config.get("db_host", ""),
        "port": int(config.get("port", 5432)),
        "database": config.get("database", ""),
        "username": config.get("username", ""),
        "password": config.get("password", ""),
        "action": AgentConstants.INSERT
    }

    try:
        db_conn_insert = DbConn(json.dumps(db_net_value_insert))
        insert_req = {
            "query_key": "INSERT_FIN_STMT_INFO",
            "params": all_insert_params
        }
        db_conn_insert.create_request(json.dumps(insert_req))
        insert_res_str = db_conn_insert.create_response()
        insert_result = json.loads(insert_res_str)

        print(f"\n  [DB 적재 완료] fin_stmt_info 테이블에 {corp_name} 총 {len(all_insert_params)}건 데이터 삽입 성공! (결과: {insert_result})")

    except Exception as e:
        print(f"[오류] fin_stmt_info 테이블 데이터 저장 중 예외 발생: {e}")
        return False

    print("\n" + "=" * 110)
    print(f"{'사업연도':^8} | {'매출액 (원)':^25} | {'영업이익 (원)':^25} | {'당기순이익 (원)':^25}")
    print("-" * 110)

    for year in range(s_year, e_year + 1):
        summary = all_summary.get(year, {})
        revenue = summary.get("매출액", summary.get("수익(매출액)", "-"))
        op_profit = summary.get("영업이익", summary.get("영업이익(손실)", "-"))
        net_profit = summary.get("당기순이익(손실)", summary.get("당기순이익", "-"))
        print(f"{year:^8} | {revenue:>25} | {op_profit:>25} | {net_profit:>25}")

    print("=" * 110 + "\n")
    return True


def process_excel_stocks(
    excel_file_path: str, 
    start_year: Optional[int] = None, 
    end_year: Optional[int] = None
):
    """
    src와 같은 레벨(프로젝트 최상단) 위치의 기업 목록 엑셀 파일에서 기업 목록을 읽어와,
    전 종목에 대해 지정한 기간의 정기보고서 재무제표 데이터를 fin_stmt_info 테이블에 적재합니다.
    """
    try:
        import openpyxl
    except ImportError:
        print("[오류] openpyxl 패키지가 설치되어 있지 않습니다. (pip install openpyxl)")
        return

    def_start, def_end = get_default_years()
    s_year = start_year if start_year is not None else def_start
    e_year = end_year if end_year is not None else def_end

    # 엑셀 파일 탐색 1순위: src와 같은 레벨인 project_root 디렉터리
    candidate_paths = [
        os.path.abspath(os.path.join(project_root, excel_file_path)),
        os.path.abspath(os.path.join(src_dir, excel_file_path)),
        os.path.abspath(excel_file_path)
    ]

    full_excel_path = None
    for path in candidate_paths:
        if os.path.exists(path):
            full_excel_path = path
            break

    if not full_excel_path:
        print(f"\n[오류] 입력하신 엑셀 파일을 찾을 수 없습니다: '{excel_file_path}'")
        print(f"       탐색한 경로 목록:")
        for p in candidate_paths:
            print(f"         - {p}")
        return

    wb = openpyxl.load_workbook(full_excel_path, data_only=True)
    ws = wb.active

    stock_list = []
    for r in range(2, ws.max_row + 1):
        raw_code = ws.cell(r, 1).value
        raw_name = ws.cell(r, 2).value
        if not raw_code or not raw_name:
            continue
        clean_code = str(raw_code).replace('\xa0', '').strip().zfill(6)
        clean_name = str(raw_name).strip()
        stock_list.append((clean_code, clean_name))

    wb.close()

    print("\n" + "=" * 95)
    print(f"        [ 기업 목록 엑셀({os.path.basename(full_excel_path)}) ({len(stock_list)}개 기업) {s_year}~{e_year}년 DB 적재 시작 ]")
    print("=" * 95)

    config = AgentUtils.load_config("agent_key.json")
    if not isinstance(config, dict) or not config:
        print("[오류] agent_key.json 설정 정보를 로드할 수 없습니다.")
        return

    db_net_value_select = {
        "host": config.get("db_host", ""),
        "port": int(config.get("port", 5432)),
        "database": config.get("database", ""),
        "username": config.get("username", ""),
        "password": config.get("password", ""),
        "action": AgentConstants.SELECT
    }

    success_count = 0
    fail_count = 0

    for idx, (stock_code, corp_name) in enumerate(stock_list, start=1):
        print(f"\n[{idx}/{len(stock_list)}] 종목코드: {stock_code} | 기업명: {corp_name} 처리 중...")
        
        target_rule_no = None
        target_corp_code = None

        try:
            db_conn_select = DbConn(json.dumps(db_net_value_select))
            select_req = {
                "query_key": "SELECT_CORP_CODE_BY_STOCK_CODE",
                "params": {"stock_code": stock_code}
            }
            db_conn_select.create_request(json.dumps(select_req))
            select_res_str = db_conn_select.create_response()
            select_rows = json.loads(select_res_str)

            if select_rows and select_rows[0].get("rule_no"):
                target_rule_no = select_rows[0]["rule_no"]
                target_corp_code = select_rows[0].get("corp_code")

        except Exception as e:
            print(f"  [경고] stock_code 조회 중 예외: {e}")

        if not target_rule_no:
            try:
                db_conn_select = DbConn(json.dumps(db_net_value_select))
                select_req = {
                    "query_key": "SELECT_CORP_CODE_BY_NAME",
                    "params": {"corp_name": corp_name}
                }
                db_conn_select.create_request(json.dumps(select_req))
                select_res_str = db_conn_select.create_response()
                select_rows = json.loads(select_res_str)

                if select_rows and select_rows[0].get("rule_no"):
                    target_rule_no = select_rows[0]["rule_no"]
                    target_corp_code = select_rows[0].get("corp_code")
            except Exception as e:
                print(f"  [경고] corp_name 조회 중 예외: {e}")

        if not target_rule_no or not target_corp_code:
            print(f"  [실패] DB 'corp_code_info'에서 종목({stock_code} / {corp_name}) 정보 매핑 불가")
            fail_count += 1
            continue

        try:
            ok = save_company_financial_statements_to_db(
                corp_name=corp_name,
                corp_code=target_corp_code,
                start_year=s_year,
                end_year=e_year
            )
            if ok:
                success_count += 1
            else:
                fail_count += 1
        except Exception as e:
            print(f"  [오류] {corp_name} 수집/적재 작업 중 오류 발생: {e}")
            fail_count += 1

    print("\n" + "=" * 95)
    print(f"        [ 기업 목록 엑셀 재무제표 DB 적재 작업 완료 ]")
    print(f"        - 성공: {success_count}개 기업 / 실패: {fail_count}개 기업 (총 {len(stock_list)}개)")
    print("=" * 95 + "\n")


def prompt_year_range():
    """사용자에게 시작연도 및 종료연도를 입력받는 도우미 함수"""
    def_start, def_end = get_default_years()
    print(f"\n[기간 설정] 기본 수집 기간: 최근 10년 ({def_start}~{def_end}년)")
    
    start_input = input(f"▶ 시작 연도 입력 (기본값 {def_start}): ").strip()
    end_input = input(f"▶ 종료 연도 입력 (기본값 {def_end}): ").strip()

    start_year = int(start_input) if start_input.isdigit() else def_start
    end_year = int(end_input) if end_input.isdigit() else def_end

    if start_year > end_year:
        print(f"[경고] 시작연도가 종료연도보다 큽니다. 기본값({def_start}~{def_end}년)으로 재설정합니다.")
        return def_start, def_end

    return start_year, end_year


def main():
    def_start, def_end = get_default_years()
    print("\n" + "=" * 55)
    print("        [ 재무제표 DB 적재 스크립트 (fin_stmt_loader) ]")
    print("=" * 55)
    print("  1. 개별 기업 재무제표 DB 적재")
    print("  2. 기업 목록 엑셀 파일(src와 같은 레벨) 전체 DB 적재")
    print("=" * 55)
    choice = input("원하는 작업 번호를 선택하세요 (1/2): ").strip()

    if choice == "1":
        corp_input = input("\n▶ 적재할 기업명을 입력하세요: ").strip()
        if not corp_input:
            print("[오류] 기업명을 입력해야 합니다.")
            return
        s_year, e_year = prompt_year_range()
        save_company_financial_statements_to_db(corp_input, start_year=s_year, end_year=e_year)
    elif choice == "2":
        excel_path = input("\n▶ 적재할 기업 목록 엑셀 파일명/경로를 입력하세요: ").strip()
        if not excel_path:
            print("[오류] 엑셀 파일명을 입력해야 합니다.")
            return
        s_year, e_year = prompt_year_range()
        process_excel_stocks(excel_file_path=excel_path, start_year=s_year, end_year=e_year)
    else:
        print("\n[오류] 잘못된 선택입니다. 1 또는 2를 입력하세요.")


if __name__ == "__main__":
    main()
