import sys
import os
import io
import json
from typing import Dict, Any

# Windows CP949 콘솔 환경에서 유니코드 출력 깨짐 방지
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Add 'src' directory to Python search path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../src')))

import services.opendart_service
from services.service_factory import ServiceFactory
from common.constants import AgentConstants
from common.utils import AgentUtils
from common.db_conn import DbConn


def clean_decimal_amount(val_str) -> str:
    """
    OpenDART 수치 문자열을 Decimal 정수/소수 수치 문자열로 변환합니다. (numeric 컬럼 저장용)
    """
    if val_str is None:
        return "0"
    cleaned = str(val_str).replace(",", "").strip()
    if not cleaned or cleaned in ("-", "None", "null"):
        return "0"
    try:
        from decimal import Decimal
        return str(Decimal(cleaned))
    except Exception:
        return "0"


def clean_int_ord(val_str) -> int:
    """
    OpenDART 'ord' 표시 순서 문자열을 int 정수로 변환합니다. (integer 컬럼 저장용)
    """
    if val_str is None:
        return 0
    cleaned = str(val_str).strip()
    if not cleaned or not cleaned.isdigit():
        return 0
    try:
        return int(cleaned)
    except Exception:
        return 0


def save_company_financial_statements_to_db(corp_name: str, corp_code: str = None, start_year: int = 2016, end_year: int = 2025):
    """
    OpendartService를 통해 지정 기업의 10년치 정기보고서 재무제표를 조회하고,
    DbConn 및 AgentUtils.get_rule_no()를 활용하여 DB의 'fin_stmt_info' 테이블에 파라미터를 바인딩하여 저장합니다.
    """
    config = AgentUtils.load_config("agent_key.json")
    if not isinstance(config, dict) or not config:
        print("[오류] agent_key.json 설정 정보를 로드할 수 없습니다.")
        return

    crtfc_key = config.get("dart_crtfc_key", "")
    if not crtfc_key:
        print("[오류] OpenDART 인증키(dart_crtfc_key)가 설정되어 있지 않습니다.")
        return

    opendart_service = ServiceFactory.create(AgentConstants.OPENDART)
    if not opendart_service:
        print("[오류] OPENDART 서비스를 초기화할 수 없습니다.")
        return

    # 1. DbConn을 통해 'corp_code_info'에서 기업의 PK (rule_no) 및 DART corp_code 조회
    db_net_value_select = {
        "host": config.get("db_host", ""),
        "port": int(config.get("port", 5432)),
        "database": config.get("database", ""),
        "username": config.get("username", ""),
        "password": config.get("password", ""),
        "action": AgentConstants.SELECT
    }

    try:
        db_conn_select = DbConn(json.dumps(db_net_value_select))
        select_req = {
            "query_key": "SELECT_CORP_CODE_BY_NAME",
            "params": {"corp_name": corp_name}
        }
        db_conn_select.create_request(json.dumps(select_req))
        select_res_str = db_conn_select.create_response()
        select_rows = json.loads(select_res_str)

        if not select_rows or not select_rows[0].get("rule_no"):
            print(f"[오류] corp_code_info에서 '{corp_name}'의 PK (rule_no)를 찾을 수 없습니다.")
            return

        fk_rule_no = select_rows[0]["rule_no"]
        db_corp_code = select_rows[0].get("corp_code", "")
        target_corp_code = corp_code or db_corp_code

        print(f"  [확인] '{corp_name}' FK (corp_code_info.rule_no): {fk_rule_no}, DART corp_code: {target_corp_code}")

    except Exception as e:
        print(f"[오류] DB에서 '{corp_name}' FK 조회 실패: {e}")
        return

    # 2. 기존 DB 데이터 중 해당 기업(corp_no)의 fin_stmt_info 데이터 삭제 (중복 저장 방지)
    db_net_value_delete = {
        "host": config.get("db_host", ""),
        "port": int(config.get("port", 5432)),
        "database": config.get("database", ""),
        "username": config.get("username", ""),
        "password": config.get("password", ""),
        "action": AgentConstants.DELETE
    }
    try:
        db_conn_delete = DbConn(json.dumps(db_net_value_delete))
        delete_req = {
            "query_key": "DELETE_FIN_STMT_INFO_BY_CORP_NO",
            "params": {"corp_no": fk_rule_no}
        }
        db_conn_delete.create_request(json.dumps(delete_req))
        del_res_str = db_conn_delete.create_response()
        print(f"  [기존 데이터 삭제] '{corp_name}' 기존 fin_stmt_info 레코드 정리 완료: {del_res_str}")
    except Exception as e:
        print(f"[경고] 기존 데이터 삭제 처리 실패: {e}")

    # 3. OpenDART 10년치 정기보고서(사업보고서 11011) 데이터 수집 및 bsns_year + account_nm 중복 방지
    print("\n" + "=" * 95)
    print(f"        [ {corp_name} 10년치({start_year}~{end_year}년) 재무제표 DB(fin_stmt_info) 중복 제거 적재 ]")
    print("=" * 95)

    reprt_code = "11011"
    all_insert_params = []
    all_summary = {}
    seen_keys = set()  # (bsns_year, account_nm) 중복 방지 집합

    for year in range(start_year, end_year + 1):
        target_fs_div = "CFS"
        price_request = {
            "crtfc_key": crtfc_key,
            "corp_code": target_corp_code,
            "bsns_year": str(year),
            "reprt_code": reprt_code,
            "fs_div": target_fs_div
        }

        try:
            res_json = opendart_service.get_financial_statements_to_json(json.dumps(price_request))
            status = res_json.get("status")
            message = res_json.get("message")

            # CFS 결과가 없거나 실패 시 OFS로 재시도
            if status != "000" or not res_json.get("list"):
                target_fs_div = "OFS"
                price_request["fs_div"] = target_fs_div
                res_json = opendart_service.get_financial_statements_to_json(json.dumps(price_request))
                status = res_json.get("status")
                message = res_json.get("message")

            if status == "000":
                items = res_json.get("list", [])
                summary = {}
                added_count = 0

                for item in items:
                    bsns_year = str(item.get("bsns_year", "")).strip()
                    account_nm = str(item.get("account_nm", "")).strip()
                    dedup_key = (bsns_year, account_nm)

                    # bsns_year + account_nm 중복 제거 체크
                    if dedup_key in seen_keys:
                        continue
                    seen_keys.add(dedup_key)

                    dec_amount = clean_decimal_amount(item.get("thstrm_amount", "0"))

                    # fin_stmt_info 저장용 파라미터 구성 (account_id, account_detail, decimal thstrm_amount 포함)
                    param = {
                        "rule_no": AgentUtils.get_rule_no(),           # PK: AgentUtils.get_rule_no() 사용
                        "corp_no": fk_rule_no,                          # FK: corp_code_info의 PK값
                        "bsns_year": bsns_year,
                        "account_id": str(item.get("account_id", "")),
                        "account_nm": account_nm,
                        "account_detail": str(item.get("account_detail", "")),
                        "thstrm_amount": dec_amount,
                        "ord": clean_int_ord(item.get("ord", 0)),
                        "thstrm_nm": str(item.get("thstrm_nm", "")),
                        "fs_div": str(item.get("fs_div", target_fs_div)),
                        "fs_nm": str(item.get("fs_nm", "연결재무제표" if target_fs_div == "CFS" else "재무제표")),
                        "sj_div": str(item.get("sj_div", "")),
                        "sj_nm": str(item.get("sj_nm", "")),
                        "rcept_no": str(item.get("rcept_no", "")),
                        "reprt_code": str(item.get("reprt_code", ""))
                    }
                    all_insert_params.append(param)
                    added_count += 1

                    summary[account_nm] = dec_amount

                all_summary[year] = summary
                print(f"  [성공] {year}년 사업보고서 ({target_fs_div}: 원본 {len(items)}개 중 bsns_year+account_nm 중복제거 후 {added_count}개 반영)")
            else:
                print(f"  [실패] {year}년 사업보고서 조회 실패 (상태코드: {status}, 메시지: {message})")

        except Exception as e:
            print(f"  [오류] {year}년 사업보고서 수집 중 예외 발생: {e}")

    # 4. DbConn을 활용하여 fin_stmt_info에 바인딩 데이터 일괄 삽입
    if not all_insert_params:
        print("[경고] 저장할 재무제표 데이터가 없습니다.")
        return

    db_net_value_insert = {
        "host": config.get("db_host", ""),
        "port": int(config.get("port", 5432)),
        "database": config.get("database", ""),
        "username": config.get("username", ""),
        "password": config.get("password", ""),
        "action": AgentConstants.INSERT
    }

    try:
        db_conn_insert = DbConn(json.dumps(db_net_value_insert))
        insert_req = {
            "query_key": "INSERT_FIN_STMT_INFO",
            "params": all_insert_params
        }
        db_conn_insert.create_request(json.dumps(insert_req))
        insert_res_str = db_conn_insert.create_response()
        insert_result = json.loads(insert_res_str)

        print(f"\n  [DB 적재 완료] fin_stmt_info 테이블에 {corp_name} 총 {len(all_insert_params)}건 데이터 삽입 성공! (결과: {insert_result})")

    except Exception as e:
        print(f"[오류] fin_stmt_info 테이블 데이터 저장 중 예외 발생: {e}")

    print("\n" + "=" * 110)
    print(f"{'사업연도':^8} | {'매출액 (원)':^25} | {'영업이익 (원)':^25} | {'당기순이익 (원)':^25}")
    print("-" * 110)

    for year in range(start_year, end_year + 1):
        summary = all_summary.get(year, {})
        revenue = summary.get("매출액", summary.get("수익(매출액)", "-"))
        op_profit = summary.get("영업이익", summary.get("영업이익(손실)", "-"))
        net_profit = summary.get("당기순이익(손실)", summary.get("당기순이익", "-"))
        print(f"{year:^8} | {revenue:>25} | {op_profit:>25} | {net_profit:>25}")

    print("=" * 110 + "\n")


def process_roe_excel_stocks(excel_file_path: str = "ROE_251124.xlsx", start_year: int = 2016, end_year: int = 2025):
    """
    ROE_251124.xlsx 엑셀 파일에서 종목 코드를 읽어와,
    해당하는 전 종목에 대해 10년치 정기보고서 재무제표 데이터를 수집 후
    DbConn을 이용해 fin_stmt_info 테이블에 (bsns_year + account_nm) 중복 제거 및 Decimal 변환 적재를 진행합니다.
    """
    try:
        import openpyxl
    except ImportError:
        print("[오류] openpyxl 패키지가 설치되어 있지 않습니다.")
        return

    full_excel_path = os.path.abspath(excel_file_path)
    if not os.path.exists(full_excel_path):
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        full_excel_path = os.path.join(project_root, excel_file_path)

    if not os.path.exists(full_excel_path):
        print(f"[오류] 엑셀 파일이 존재하지 않습니다: {full_excel_path}")
        return

    wb = openpyxl.load_workbook(full_excel_path, data_only=True)
    ws = wb.active

    stock_list = []
    for r in range(2, ws.max_row + 1):
        raw_code = ws.cell(r, 1).value
        raw_name = ws.cell(r, 2).value
        if not raw_code or not raw_name:
            continue
        clean_code = str(raw_code).replace('\xa0', '').strip().zfill(6)
        clean_name = str(raw_name).strip()
        stock_list.append((clean_code, clean_name))

    wb.close()

    print("\n" + "=" * 95)
    print(f"        [ ROE 엑셀({os.path.basename(full_excel_path)}) 종목 목록 ({len(stock_list)}개 기업) 10년치 DB 적재 시작 ]")
    print("=" * 95)

    config = AgentUtils.load_config("agent_key.json")
    if not isinstance(config, dict) or not config:
        print("[오류] agent_key.json 설정 정보를 로드할 수 없습니다.")
        return

    db_net_value_select = {
        "host": config.get("db_host", ""),
        "port": int(config.get("port", 5432)),
        "database": config.get("database", ""),
        "username": config.get("username", ""),
        "password": config.get("password", ""),
        "action": AgentConstants.SELECT
    }

    success_count = 0
    fail_count = 0

    for idx, (stock_code, corp_name) in enumerate(stock_list, start=1):
        print(f"\n[{idx}/{len(stock_list)}] 종목코드: {stock_code} | 기업명: {corp_name} 처리 중...")
        
        target_rule_no = None
        target_corp_code = None

        try:
            db_conn_select = DbConn(json.dumps(db_net_value_select))
            select_req = {
                "query_key": "SELECT_CORP_CODE_BY_STOCK_CODE",
                "params": {"stock_code": stock_code}
            }
            db_conn_select.create_request(json.dumps(select_req))
            select_res_str = db_conn_select.create_response()
            select_rows = json.loads(select_res_str)

            if select_rows and select_rows[0].get("rule_no"):
                target_rule_no = select_rows[0]["rule_no"]
                target_corp_code = select_rows[0].get("corp_code")

        except Exception as e:
            print(f"  [경고] stock_code 조회 중 예외: {e}")

        if not target_rule_no:
            try:
                db_conn_select = DbConn(json.dumps(db_net_value_select))
                select_req = {
                    "query_key": "SELECT_CORP_CODE_BY_NAME",
                    "params": {"corp_name": corp_name}
                }
                db_conn_select.create_request(json.dumps(select_req))
                select_res_str = db_conn_select.create_response()
                select_rows = json.loads(select_res_str)

                if select_rows and select_rows[0].get("rule_no"):
                    target_rule_no = select_rows[0]["rule_no"]
                    target_corp_code = select_rows[0].get("corp_code")
            except Exception as e:
                print(f"  [경고] corp_name 조회 중 예외: {e}")

        if not target_rule_no or not target_corp_code:
            print(f"  [실패] DB 'corp_code_info'에서 종목({stock_code} / {corp_name}) 정보 매핑 불가")
            fail_count += 1
            continue

        try:
            save_company_financial_statements_to_db(
                corp_name=corp_name,
                corp_code=target_corp_code,
                start_year=start_year,
                end_year=end_year
            )
            success_count += 1
        except Exception as e:
            print(f"  [오류] {corp_name} 수집/적재 작업 중 오류 발생: {e}")
            fail_count += 1

    print("\n" + "=" * 95)
    print(f"        [ ROE 엑셀 전종목 10년치 재무제표 DB 적재 작업 완료 ]")
    print(f"        - 성공: {success_count}개 기업 / 실패: {fail_count}개 기업 (총 {len(stock_list)}개)")
    print("=" * 95 + "\n")


def main():
    # ROE_251124.xlsx 엑셀 내 73개 종목 전체 10년치(2016~2025년) 정기보고서(사업보고서) 재무제표 DB(fin_stmt_info) 적재 실행
    process_roe_excel_stocks("ROE_251124.xlsx", start_year=2016, end_year=2025)


if __name__ == "__main__":
    main()